"""Parse the 初期診療・救急科 student roster Excel and report this week's group.

Usage:
    python roster.py --date 2026-05-21 [--json]
    python roster.py --week 16 [--json]

The Excel layout (sheet "診療科<評価票>", headers at row 7, data from row 8):
    Col A 'キー'         - 班番号
    Col B 'G'            - 班番号 (duplicate of A)
    Col C '開始日'       - rotation start date
    Col D '終了日'       - rotation end date
    Col E '学生番号'     - long student ID (not used)
    Col F '学生氏名'     - kanji name (used for acknowledgment salutation
                          and as a name-based fallback when a student submits
                          from a personal address that is not in the roster)
    Col J 'メール'       - student email (used for Gmail search)
    Col K 'PHS'          - 学籍番号 (4-5 digit short ID — used for filenames)

We map 開始日 → 年度 (Japanese academic year, April–March):
    開始日 <  2026-04-01  →  '2025年度'
    開始日 >= 2026-04-01  →  '2026年度'
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SHEET = "診療科<評価票> "  # NOTE: trailing space in the actual sheet name
HEADER_ROW = 7
COL_KEY = 1     # A: 班番号
COL_START = 3   # C: 開始日
COL_END = 4     # D: 終了日
COL_NAME = 6    # F: 学生氏名 (kanji)
COL_EMAIL = 10  # J: メール
COL_PHS = 11    # K: PHS = 学籍番号


@dataclass(frozen=True)
class Student:
    phs: str          # 学籍番号 (e.g. "6193")
    name: str         # 学生氏名 (kanji, e.g. "塙純")
    email: str        # student email
    group: int        # 班番号


@dataclass(frozen=True)
class Group:
    group: int                # 班番号
    start: date               # 開始日
    end: date                 # 終了日
    fiscal_year: str          # "2025年度" or "2026年度"
    students: list[Student]


def fiscal_year_of(d: date) -> str:
    """Japanese academic year — starts April 1."""
    return f"{d.year - 1}年度" if d.month < 4 else f"{d.year}年度"


def load_groups(xlsx_path: Path) -> list[Group]:
    """Parse all groups from the roster. Returns groups sorted by start date."""
    wb = load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET!r} not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]

    rows_by_group: dict[int, list[Student]] = {}
    group_dates: dict[int, tuple[date, date]] = {}

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        grp = ws.cell(row=row, column=COL_KEY).value
        start = ws.cell(row=row, column=COL_START).value
        end = ws.cell(row=row, column=COL_END).value
        name = ws.cell(row=row, column=COL_NAME).value
        email = ws.cell(row=row, column=COL_EMAIL).value
        phs = ws.cell(row=row, column=COL_PHS).value

        if grp is None or start is None or end is None:
            continue
        if not isinstance(start, datetime) or not isinstance(end, datetime):
            continue
        if not isinstance(email, str) or "@" not in email:
            logger.warning("Row %d: missing/invalid email for group %s", row, grp)
            continue

        # Normalize the kanji name (strip whitespace; keep internal spaces as-is so
        # that names like "塙 純" round-trip faithfully). Tolerate missing names.
        name_str = str(name).strip() if name is not None else ""

        start_d, end_d = start.date(), end.date()
        rows_by_group.setdefault(int(grp), []).append(
            Student(phs=str(phs), name=name_str, email=email.strip(), group=int(grp))
        )
        group_dates[int(grp)] = (start_d, end_d)

    groups: list[Group] = []
    for g, students in rows_by_group.items():
        s, e = group_dates[g]
        groups.append(Group(group=g, start=s, end=e, fiscal_year=fiscal_year_of(s), students=students))
    return sorted(groups, key=lambda x: x.start)


def find_current_group(groups: list[Group], today: date, grace_days: int = 3) -> Group | None:
    """Find the group whose rotation window covers `today` (with grace days for late submissions).

    A submission window is [開始日, 終了日 + grace_days]. If `today` falls in multiple
    windows (overlap), prefer the one whose 開始日 is most recent (i.e., currently running).
    """
    candidates = [g for g in groups if g.start <= today <= g.end + timedelta(days=grace_days)]
    if not candidates:
        return None
    return max(candidates, key=lambda g: g.start)


def find_group_by_number(groups: list[Group], group_number: int) -> Group | None:
    for g in groups:
        if g.group == group_number:
            return g
    return None


def to_dict(g: Group) -> dict:
    d = asdict(g)
    d["start"] = g.start.isoformat()
    d["end"] = g.end.isoformat()
    return d


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Path to roster .xlsx (default: $ROSTER_XLSX env var)")
    parser.add_argument("--date", type=str, default=None,
                        help="Reference date YYYY-MM-DD (default: today)")
    parser.add_argument("--week", type=int, default=None,
                        help="Look up by 班番号 instead of date")
    parser.add_argument("--json", action="store_true", help="Emit JSON to stdout")
    parser.add_argument("--grace-days", type=int, default=3,
                        help="Days past 終了日 still considered 'this week'")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    import os
    xlsx_path = args.xlsx or (Path(os.environ["ROSTER_XLSX"]) if "ROSTER_XLSX" in os.environ else None)
    if xlsx_path is None:
        print("ERROR: provide --xlsx or set $ROSTER_XLSX in environment", file=sys.stderr)
        return 2

    groups = load_groups(xlsx_path)

    if args.week is not None:
        target = find_group_by_number(groups, args.week)
        if target is None:
            print(f"ERROR: 班番号 {args.week} not in roster", file=sys.stderr)
            return 1
    else:
        ref = date.fromisoformat(args.date) if args.date else date.today()
        target = find_current_group(groups, ref, grace_days=args.grace_days)
        if target is None:
            print(f"INFO: no rotation group active for {ref}", file=sys.stderr)
            return 0  # not an error — just a quiet week

    payload = {
        "week": target.group,
        "fiscal_year": target.fiscal_year,
        "start": target.start.isoformat(),
        "end": target.end.isoformat(),
        "rdm_folder": f"ポリクリ{target.group}班",
        "students": [
            {"phs": s.phs, "name": s.name, "email": s.email} for s in target.students
        ],
    }

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        print(f"Group {target.group} | {target.fiscal_year} | {target.start} → {target.end}")
        print(f"  RDM folder: {payload['rdm_folder']}")
        print(f"  Students ({len(target.students)}):")
        for s in target.students:
            print(f"    - {s.phs}  {s.name}  {s.email}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
